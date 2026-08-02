import sys
from pathlib import Path
p = Path(r"sql/29.07 EMO - CONSORCIO LOS ANDES(REG) (1).xlsx")
print('exists', p.exists(), 'size', p.stat().st_size if p.exists() else -1)
try:
    import openpyxl
except Exception as e:
    print('openpyxl_missing', e)
    raise
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb.active
print('sheet', ws.title, 'rows', ws.max_row, 'cols', ws.max_column)
for r in [1,2,3]:
    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
    print('row', r, vals)
